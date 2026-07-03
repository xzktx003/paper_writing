"""xlsx_aggregate_consistency.py — cross-XLSX same-quantity sum / row consistency.

A deterministic, single-purpose utility script. No LLM SDK. No "skeleton -> enrich"
two-stage orchestration. Use during the numerical-evidence track to catch
**inter-XLSX inconsistency** — two source-data files in the same paper that
purport to carry the same aggregate quantity but disagree.

Usage:
    python xlsx_aggregate_consistency.py <xlsx-files>... [--tol 0.001] [--entity-overlap 0.5]

What it does:
    1. Loads every XLSX, every sheet, every numeric column with its (header, unit
       family, unit factor, entity_column, entity_value_set, value_sum, value_count).
    2. For every pair of columns from DIFFERENT (file, sheet) origins:
       - Skip if either column has no recognized unit, or units belong to different
         families, or the unit factor differs (different scale — different quantity).
       - Skip if entity-value-set overlap is below --entity-overlap (default 50%).
       - Compute sum_a vs sum_b. If |Δ|/max > --tol but the overlap is broad, FLAG
         the pair as a "same-quantity inconsistency".
    3. For every flagged pair, also compute per-row deltas across the overlapping
       entities and report sign-consistency: if all deltas share the same sign, the
       mismatch is systematic (scoping or pipeline difference); if mixed, it's
       likely per-row noise or rounding.

Why this exists (empirical baseline):
    Hu et al. 2026 Nature paper (10.1038/s41586-026-10570-z) supplement contained
    two source-data tables (MOESM3 Fig1c, MOESM6 Fig4a) that both reported per-
    province annual solar+wind generation for 2022 China. Their sums differed by
    6.94 TWh (0.62%) with ALL 31 provinces showing the same sign of deviation
    (Fig1c > Fig4a). The audit flagged this as Level 1; this sweeper detects the
    class automatically.

    The pattern is generic across Nature / Cell / sub-journal papers wherever:
      - Multiple source-data tables cover overlapping entity sets (gene panels,
        sample cohorts, geographic regions, etc.)
      - The same metric is reported in two contexts (e.g., 'summary table' and
        'figure source data')
      - A scoping or filtering difference between the analysis pipelines produces
        small systematic per-entity discrepancies

Limitations (real, document them):
    - Entity-column auto-detection picks the column with most unique non-numeric
      values. Override with --entity-col file:sheet:colN if auto-detect fails.
    - Currency columns are reported (family=currency) but factor cannot resolve
      cross-currency; the script does not attempt FX conversion.
    - Pairs from the SAME (file, sheet) are not compared (within-sheet consistency
      is the job of decimal_match.py and the deterministic-column-pair sweep).

This script is intentionally < 300 lines.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
import re
from typing import Optional

try:
    import openpyxl
except ImportError:
    sys.stderr.write("install openpyxl first: pip install openpyxl\n")
    sys.exit(2)

# Reuse the unit taxonomy from magnitude_consistency.py
_TOOLS_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(_TOOLS_DIR))
try:
    from magnitude_consistency import (  # noqa: E402  # type: ignore
        unit_family, unit_to_canonical, header_unit,
    )
except ImportError as exc:
    sys.stderr.write(f"could not import magnitude_consistency (must sit next to this script): {exc}\n")
    sys.exit(2)


def auto_detect_entity_col(ws) -> Optional[int]:
    """Pick the column with most unique non-numeric values as the entity column."""
    n_rows = ws.max_row - 1
    if n_rows < 5:
        return None
    best_col, best_unique = None, 0
    for c in range(1, ws.max_column + 1):
        non_num = []
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip():
                non_num.append(v.strip())
        if len(non_num) < n_rows * 0.5:
            continue
        u = len(set(non_num))
        if u > best_unique:
            best_unique, best_col = u, c
    return best_col


def column_signature(ws, col: int, entity_col: Optional[int]) -> dict:
    """Build the per-column signature: header, unit family, sum, entity-keyed map."""
    header = ws.cell(1, col).value
    header_str = str(header) if header is not None else ""
    hunit = header_unit(header_str)
    fam = unit_family(hunit) if hunit else None
    factor = unit_to_canonical(hunit) if hunit else None
    # Entity -> value map (last write wins if duplicate entities)
    entity_to_value: dict[str, float] = {}
    total = 0.0
    count = 0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col).value
        if not isinstance(v, (int, float)) or isinstance(v, bool):
            continue
        ent = ws.cell(r, entity_col).value if entity_col else None
        ent_key = str(ent).strip() if isinstance(ent, str) and ent.strip() else f"_row{r}"
        entity_to_value[ent_key] = float(v)
        total += float(v)
        count += 1
    return {
        "header": header_str,
        "unit": hunit,
        "family": fam,
        "factor": factor,
        "entity_to_value": entity_to_value,
        "sum": total,
        "count": count,
    }


def load_all_columns(xlsx_paths: list[str]) -> list[dict]:
    """Return one signature dict per (file, sheet, numeric column)."""
    out: list[dict] = []
    for path in xlsx_paths:
        p = Path(path)
        if not p.is_file():
            sys.stderr.write(f"not a file: {path}\n")
            continue
        try:
            wb = openpyxl.load_workbook(p, data_only=True)
        except Exception as exc:
            sys.stderr.write(f"failed to open {path}: {exc}\n")
            continue
        for sn in wb.sheetnames:
            ws = wb[sn]
            if ws.max_row < 2:
                continue
            ent_col = auto_detect_entity_col(ws)
            for c in range(1, ws.max_column + 1):
                if c == ent_col:
                    continue
                sig = column_signature(ws, c, ent_col)
                if sig["count"] < 3:
                    continue  # ignore tiny columns
                if sig["family"] is None:
                    continue  # unit-less columns — cannot reason about
                sig["file"] = p.name
                sig["sheet"] = sn
                sig["col"] = c
                sig["entity_col"] = ent_col
                out.append(sig)
    return out


def entity_overlap(a: set, b: set) -> float:
    """Overlap as |a ∩ b| / max(|a|, |b|). 1.0 = identical entity sets."""
    if not a or not b:
        return 0.0
    return len(a & b) / max(len(a), len(b))


# Common unit suffixes / SI prefixes / generic stopwords to strip from header tokens
_HEADER_STOP = {
    "twh", "gwh", "mwh", "kwh", "wh", "tw", "gw", "mw", "kw", "w",
    "tj", "gj", "mj", "j", "kg", "mg", "ng", "ug", "g", "t",
    "mol", "mm", "um", "nm", "km", "m", "cm", "mm2", "km2", "m2", "ha", "l", "ml",
    "hz", "mhz", "ghz", "khz", "pa", "kpa", "mpa", "v", "kv", "mv", "a", "ka",
    "bp", "kb", "mb", "tco2", "co2",
    "annual", "total", "value", "sum", "per", "in", "of", "and", "the", "a", "an",
    "id", "name", "year", "yr",
}


def header_semantic_tokens(header: str) -> set[str]:
    """Tokens from header excluding unit suffixes and generic words.

    Splits on underscore / camelCase / whitespace / parens.
    """
    if not header:
        return set()
    # snake_case, kebab-case, spaces, parens
    raw = re.split(r"[_\s\-\(\)/,]+", header)
    # camelCase split
    expanded: list[str] = []
    for part in raw:
        if not part:
            continue
        # Split camelCase to "annual" + "Solar" etc.
        camel_parts = re.findall(r"[A-Z]?[a-z]+|[A-Z]+(?=[A-Z]|$)", part)
        if camel_parts:
            expanded.extend(camel_parts)
        else:
            expanded.append(part)
    toks = {t.lower().strip() for t in expanded if t and t.strip()}
    return {t for t in toks if t and t not in _HEADER_STOP and len(t) >= 2}


def main() -> int:
    ap = argparse.ArgumentParser(description=(__doc__ or "").split("\n\n")[0])
    ap.add_argument("files", nargs="+", help="XLSX files to cross-compare")
    ap.add_argument("--tol-low", type=float, default=0.001,
                    help="lower bound of relative-diff to flag (default 0.001 = 0.1 percent). "
                         "Pairs closer than this are exact-match — no finding.")
    ap.add_argument("--tol-high", type=float, default=0.05,
                    help="upper bound of relative-diff to flag (default 0.05 = 5 percent). "
                         "Pairs further apart than this are likely different metrics, not the same quantity.")
    ap.add_argument("--entity-overlap", type=float, default=0.5,
                    help="require >= this fraction of entity overlap between two columns (default 0.5)")
    ap.add_argument("--header-overlap", type=int, default=1,
                    help="require >= this many shared NON-UNIT tokens between the two column headers "
                         "(default 1; set 0 to disable the semantic-name check)")
    ap.add_argument("--min-rows", type=int, default=5,
                    help="skip columns with fewer than this many numeric rows (default 5)")
    ap.add_argument("--max-pairs", type=int, default=40,
                    help="report at most this many pairs (default 40)")
    args = ap.parse_args()

    cols = load_all_columns(args.files)
    cols = [c for c in cols if c["count"] >= args.min_rows]
    sys.stderr.write(f"loaded {len(cols)} numeric columns with recognized units from {len(args.files)} file(s)\n")

    # Build pairwise comparisons; skip same (file, sheet) pairs and family/factor mismatches.
    flags = []
    seen_pair = set()
    for i in range(len(cols)):
        for j in range(i + 1, len(cols)):
            a, b = cols[i], cols[j]
            if (a["file"], a["sheet"]) == (b["file"], b["sheet"]):
                continue
            if a["family"] != b["family"]:
                continue
            if a["factor"] != b["factor"]:
                continue  # different power-of-10 scale = different quantity
            # Skip currency family (no factor comparison)
            if a["family"] == "currency":
                continue
            overlap = entity_overlap(set(a["entity_to_value"]), set(b["entity_to_value"]))
            if overlap < args.entity_overlap:
                continue
            sum_max = max(abs(a["sum"]), abs(b["sum"]))
            if sum_max == 0:
                continue
            rel_diff = abs(a["sum"] - b["sum"]) / sum_max
            if rel_diff < args.tol_low:
                continue  # within tolerance, agree
            if rel_diff > args.tol_high:
                continue  # too large — likely different metrics, not the same quantity
            # Semantic header-overlap check: require ≥ N non-unit tokens shared.
            # Exception: if EITHER header is "generic" (no semantic tokens — only
            # contains stopwords + unit suffix, e.g. 'total_TWh' or 'annual_load_TWh'),
            # skip the check — a generic-named summary column is allowed to match
            # a descriptive column on the same family + entity set.
            if args.header_overlap > 0:
                a_toks = header_semantic_tokens(a["header"])
                b_toks = header_semantic_tokens(b["header"])
                a_generic = len(a_toks) == 0
                b_generic = len(b_toks) == 0
                if not (a_generic or b_generic):
                    if len(a_toks & b_toks) < args.header_overlap:
                        continue
            # Per-row delta sign consistency on overlap
            overlap_keys = set(a["entity_to_value"]) & set(b["entity_to_value"])
            row_deltas = [a["entity_to_value"][k] - b["entity_to_value"][k] for k in overlap_keys]
            pos = sum(1 for d in row_deltas if d > 1e-12)
            neg = sum(1 for d in row_deltas if d < -1e-12)
            same_sign = max(pos, neg) / max(1, pos + neg)  # 1.0 = all same sign
            key = (a["file"], a["sheet"], a["col"], b["file"], b["sheet"], b["col"])
            if key in seen_pair:
                continue
            seen_pair.add(key)
            flags.append({
                "a": a, "b": b, "rel_diff": rel_diff, "overlap": overlap,
                "same_sign": same_sign, "n_overlap": len(overlap_keys),
                "n_pos": pos, "n_neg": neg,
            })

    # Sort: systematic same-sign mismatches first, then by relative diff descending
    flags.sort(key=lambda f: (-f["same_sign"], -f["rel_diff"]))

    if not flags:
        sys.stderr.write("\n=== 0 flag(s) ===\n")
        return 0

    print(f"\n=== {len(flags)} pair(s) flagged ===\n")
    for f in flags[: args.max_pairs]:
        a, b = f["a"], f["b"]
        sys.stdout.write(f"*** AGGREGATE-MISMATCH ({f['rel_diff'] * 100:.3f} percent diff)")
        if f["same_sign"] >= 0.95 and f["n_overlap"] >= 5:
            sys.stdout.write("  [SYSTEMATIC — all same sign]")
        sys.stdout.write(" ***\n")
        print(f"  A: {a['file']}:{a['sheet']}!col{a['col']} header={a['header']!r} unit={a['unit']!r} "
              f"sum={a['sum']:.4f} n={a['count']}")
        print(f"  B: {b['file']}:{b['sheet']}!col{b['col']} header={b['header']!r} unit={b['unit']!r} "
              f"sum={b['sum']:.4f} n={b['count']}")
        print(f"  Δ_sum={a['sum'] - b['sum']:+.4f}  entity_overlap={f['overlap']:.1%}  "
              f"per-row: n_overlap={f['n_overlap']} pos={f['n_pos']} neg={f['n_neg']} "
              f"same_sign={f['same_sign']:.1%}")
        # Show up to 5 largest absolute deltas
        overlap_keys = set(a["entity_to_value"]) & set(b["entity_to_value"])
        row_deltas = sorted(
            ((k, a["entity_to_value"][k] - b["entity_to_value"][k]) for k in overlap_keys),
            key=lambda kv: -abs(kv[1]),
        )
        for k, d in row_deltas[:5]:
            print(f"      {k!r:>25}: A={a['entity_to_value'][k]:>10.4f}  B={b['entity_to_value'][k]:>10.4f}  Δ={d:+.4f}")
        print()
    sys.stderr.write(f"\n=== {len(flags)} pair(s) flagged ===\n")
    return 0 if not flags else 1


if __name__ == "__main__":
    sys.exit(main())
