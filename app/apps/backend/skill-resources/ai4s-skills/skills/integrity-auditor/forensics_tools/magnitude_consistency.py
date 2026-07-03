"""magnitude_consistency.py — supplement-text vs source-data XLSX unit/scale consistency.

A deterministic, single-purpose utility script. No LLM SDK. No "skeleton -> enrich"
two-stage orchestration. Use during the numerical-evidence track to catch
**unit-confusion / scale-mismatch** editorial or fabrication errors by
cross-referencing numeric claims in supplementary text against the underlying
source-data XLSX cells.

Usage:
    python magnitude_consistency.py --text supp1.txt supp2.txt --xlsx data1.xlsx data2.xlsx

What it does:
    1. From each XLSX, extracts every numeric cell with its column header. If the
       header carries a recognizable unit suffix (e.g., 'wind_generation_GWh',
       'load_TWh', 'distance_km'), the unit is recorded.
    2. From each text file, extracts every numeric claim with the unit token
       appearing in a +/- 80-char context window.
    3. For each (text_value V_t, text_unit U_t) vs (xlsx_value V_x, xlsx_unit U_x):
       - If |V_t - V_x| / V_x < 1% AND U_t and U_x are *both recognized* AND
         differ by a power-of-10 factor (e.g., TWh vs GWh = 1000), FLAG.
    4. Also reports value-proximity hits where text mentions a value matching an
       XLSX cell at exactly 10x / 100x / 1000x AND the row-entity name appears
       in the text context (entity-overlap check, language-agnostic).

Why this exists (empirical baseline):
    Hu et al. 2026 Nature paper (10.1038/s41586-026-10570-z) carried a
    systematic 1000x unit error in MOESM1 + MOESM2 supplementary text: per-county
    annual wind generation reported as 'TWh' when source-data MOESM3 was in
    GWh. The headline national totals were correct; the per-county
    supplementary tables would have implied a single Chinese county producing
    ~34% of China's national electricity. The skill's v1.5 sweepers did not catch
    this; auditor caught it by reading and recognizing the implausible
    magnitude. This sweeper mechanizes that recognition.

Limitations (real, document them):
    - Unit suffix detection from XLSX headers requires a recognizable pattern.
      Headers like 'value' or 'metric' with no unit will be skipped.
    - The entity-overlap check is token-based and language-agnostic but
      cross-language matches (Chinese county name in XLSX vs English transliteration
      in text) require a bilingual gazetteer that this tool does not ship with.
      Pass --bilingual-map FILE to inject a JSON {english: chinese} mapping.
    - The 1% value-proximity threshold is conservative; very-similar
      coincidences (e.g., 100.0 in TWh near 100.5 in GWh of unrelated rows)
      may produce false positives. Filter the report manually.

This script is intentionally < 250 lines. Pair with image_dup.py and
decimal_match.py in the forensics_tools/ pipeline.
"""

from __future__ import annotations

import argparse
import collections
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write("install openpyxl first: pip install openpyxl\n")
    sys.exit(2)


# Number pattern: 2,900.5  or  2900  or  2900.5 — not part of identifier
NUM_RE = re.compile(r"(?<![A-Za-z_\d.])(\d{1,3}(?:,\d{3})+(?:\.\d+)?|\d+\.\d+|\d{2,})(?![A-Za-z_\d])")


def parse_num(s: str) -> float:
    return float(s.replace(",", ""))


# === Unit taxonomy ===
#
# SI prefix factors. Includes the Greek µ and ASCII u as micro.
SI_PREFIXES: dict[str, float] = {
    "y": 1e-24, "z": 1e-21, "a": 1e-18, "f": 1e-15, "p": 1e-12, "n": 1e-9,
    "µ": 1e-6, "μ": 1e-6, "u": 1e-6, "c": 1e-2,  # 'm' (milli) handled specially below
    "": 1.0, "k": 1e3, "M": 1e6, "G": 1e9, "T": 1e12, "P": 1e15, "E": 1e18,
    "Z": 1e21, "Y": 1e24,
}
# 'm' is milli-, but it's also a valid base unit (metre). Tested last so longer bases win.
SI_PREFIXES_WITH_MILLI = dict(SI_PREFIXES, m=1e-3)

# Base units. Sorted by symbol length DESC to prefer longer matches first.
# Each tuple: (symbol, family, intrinsic_factor_within_family).
# `intrinsic_factor` is the multiplier from this base to the family's canonical base unit.
BASE_UNITS: list[tuple[str, str, float]] = sorted([
    # Energy (canonical = Wh)
    ("Wh", "energy", 1.0),
    ("J", "energy", 1.0 / 3600.0),       # 1 J = 1/3600 Wh
    ("cal", "energy", 4.184 / 3600.0),   # 1 cal = 4.184 J
    ("eV", "energy", 1.602e-19 / 3600.0),
    ("BTU", "energy", 1055.06 / 3600.0),
    # Power (canonical = W)
    ("W", "power", 1.0),
    # Mass (canonical = g)
    ("g", "mass", 1.0),
    ("t", "mass", 1e6),                  # 1 tonne = 1e6 g
    ("Da", "mass", 1.66054e-24),
    # Length (canonical = m)
    ("m", "length", 1.0),
    # Area (canonical = m2)
    ("m2", "area", 1.0),
    ("ha", "area", 1e4),
    # Volume (canonical = L)
    ("L", "volume", 1.0),
    ("l", "volume", 1.0),
    ("m3", "volume", 1e3),
    # Time (canonical = s)
    ("s", "time", 1.0),
    ("min", "time", 60.0),
    ("h", "time", 3600.0),
    ("d", "time", 86400.0),
    ("yr", "time", 3.1557600e7),
    ("y", "time", 3.1557600e7),
    # Voltage (canonical = V)
    ("V", "voltage", 1.0),
    # Current (canonical = A)
    ("A", "current", 1.0),
    # Frequency (canonical = Hz)
    ("Hz", "frequency", 1.0),
    # Pressure (canonical = Pa)
    ("Pa", "pressure", 1.0),
    ("bar", "pressure", 1e5),
    ("atm", "pressure", 101325.0),
    # Amount of substance (canonical = mol)
    ("mol", "amount", 1.0),
    # Concentration (canonical = M = mol/L)
    ("M", "concentration", 1.0),
    # Force (canonical = N)
    ("N", "force", 1.0),
    # Ionising-radiation dose (canonical = Gy)
    ("Gy", "dose", 1.0),
    ("Sv", "dose_eq", 1.0),
    # Genomics (canonical = bp)
    ("bp", "bp", 1.0),
    # CO2 emissions (canonical = tCO2)
    ("tCO2", "co2", 1.0),
    # Currency (canonical = USD; intrinsic factors are not maintained — currency family is here
    # only so cross-currency confusion can be detected at the family level.)
    ("USD", "currency", 1.0),
    ("EUR", "currency", 1.0),
    ("CNY", "currency", 1.0),
    ("RMB", "currency", 1.0),
    ("GBP", "currency", 1.0),
    ("JPY", "currency", 1.0),
], key=lambda x: -len(x[0]))


def parse_unit(u: str) -> tuple[str | None, float | None]:
    """Parse a unit token to (family, factor) where factor converts to the family's canonical base.

    Returns (None, None) if unrecognized. Currency family carries factor 1.0 for all
    symbols (we cannot convert between currencies without an exchange-rate table; this is
    intentional — same-family mismatch is still detectable as 'currency confusion').
    """
    if not u:
        return (None, None)
    # Try each base unit, longest-symbol-first
    for sym, family, base_factor in BASE_UNITS:
        if u.endswith(sym):
            prefix = u[: -len(sym)] if sym else u
            # Strict matching of the prefix
            if prefix in SI_PREFIXES_WITH_MILLI:
                return (family, SI_PREFIXES_WITH_MILLI[prefix] * base_factor)
    return (None, None)


def unit_family(u: str) -> str | None:
    """Convenience: just the family."""
    fam, _ = parse_unit(u)
    return fam


def unit_to_canonical(u: str) -> float | None:
    """Convenience: just the factor to canonical."""
    _, f = parse_unit(u)
    return f


def same_unit_family(a: str, b: str) -> bool:
    """True iff both units parse to the same family. Returns False on unknowns."""
    fa = unit_family(a)
    fb = unit_family(b)
    return fa is not None and fa == fb


# Regex for finding a unit token in text. Built from the BASE_UNITS list so it stays in sync.
_BASE_ALTERNATION = "|".join(re.escape(s) for s, _, _ in BASE_UNITS)
_PREFIX_ALTERNATION = "(?:y|z|a|f|p|n|µ|μ|u|m|c|k|M|G|T|P|E|Z|Y)?"
UNIT_TOKEN = r"(?:" + _PREFIX_ALTERNATION + r"(?:" + _BASE_ALTERNATION + r"))"
UNIT_CTX_RE = re.compile(r"\b(" + UNIT_TOKEN + r")\b")


def header_unit(h: str) -> str | None:
    """Extract unit from a column header like 'wind_generation_GWh' -> 'GWh'.

    Accepts:
      - trailing _UNIT (snake_case headers)
      - trailing (UNIT) parens
      - trailing UNIT preceded by space, hyphen, or comma
    Only returns the unit if it parses to a known family.
    """
    if not h:
        return None
    for pattern in (r"_(" + UNIT_TOKEN + r")$",
                    r"\((" + UNIT_TOKEN + r")\)\s*$",
                    r"[\s,\-](" + UNIT_TOKEN + r")$"):
        m = re.search(pattern, h)
        if m and unit_family(m.group(1)) is not None:
            return m.group(1)
    return None


def tokens(text: str) -> set[str]:
    """Return entity-matching tokens: lowercased Latin words (≥3 chars) + Chinese 2-8 char runs."""
    out = set()
    for w in re.findall(r"[A-Za-z][A-Za-z\-]{2,}", text):
        out.add(w.lower())
    for w in re.findall(r"[\u4e00-\u9fff]{2,8}", text):
        out.add(w)
    return out


def auto_detect_entity_col(ws) -> int | None:
    """Pick the column with most non-numeric unique values as the entity column."""
    best_col, best_unique = None, 0
    n_rows = ws.max_row - 1
    if n_rows < 5:
        return None
    for c in range(1, ws.max_column + 1):
        non_num = []
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip():
                non_num.append(v)
        if len(non_num) < n_rows * 0.5:
            continue
        u = len(set(non_num))
        if u > best_unique:
            best_unique, best_col = u, c
    return best_col


def load_xlsx(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    entries: list[dict] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row < 2:
            continue
        ent_col = auto_detect_entity_col(ws)
        headers = {c: ws.cell(1, c).value for c in range(1, ws.max_column + 1)}
        for r in range(2, ws.max_row + 1):
            ent_raw = ws.cell(r, ent_col).value if ent_col else None
            ent_tokens = tokens(str(ent_raw)) if ent_raw else set()
            for c in range(1, ws.max_column + 1):
                if c == ent_col:
                    continue
                v = ws.cell(r, c).value
                if not isinstance(v, (int, float)) or isinstance(v, bool):
                    continue
                hdr = str(headers.get(c, "")) if headers.get(c) else ""
                entries.append({
                    "file": Path(path).name,
                    "sheet": sn,
                    "coord": ws.cell(r, c).coordinate,
                    "value": float(v),
                    "header": hdr,
                    "header_unit": header_unit(hdr),
                    "entity_tokens": ent_tokens,
                    "entity_raw": str(ent_raw) if ent_raw else "",
                })
    return entries


YEAR_RE = re.compile(r"^(19|20)\d{2}$")
UNIT_RIGHT_RE = re.compile(r"^\s*(" + UNIT_TOKEN + r")\b")


def extract_text_claims(text: str, ctx_window: int = 80, unit_max_gap: int = 5) -> list[dict]:
    """Extract numeric claims with unit attached only if unit immediately follows.

    A unit is attached to a number only when the unit token appears within
    `unit_max_gap` characters to the right of the number (no intervening
    non-space punctuation that breaks the pair). This rules out:
      - year references: '2029, it delivers 43 TWh' — '2029' gets unit=None
      - cross-sentence units: '2900. The next year reports TWh values' — also None
    """
    claims = []
    for m in NUM_RE.finditer(text):
        raw = m.group(1)
        if YEAR_RE.match(raw.replace(",", "")):
            continue  # skip years
        try:
            v = parse_num(raw)
        except ValueError:
            continue
        if v < 1.0:
            continue
        s = max(0, m.start() - ctx_window)
        e = min(len(text), m.end() + ctx_window)
        ctx = text[s:e]
        # Look for unit only in the small right-adjacent window
        right_window = text[m.end(): min(len(text), m.end() + unit_max_gap)]
        unit = None
        u_match = UNIT_RIGHT_RE.match(right_window)
        if u_match:
            unit = u_match.group(1)
        claims.append({
            "value": v,
            "unit": unit,
            "context": re.sub(r"\s+", " ", ctx).strip(),
            "context_tokens": tokens(ctx),
        })
    return claims


def main() -> int:
    ap = argparse.ArgumentParser(description=(__doc__ or "").split("\n\n")[0])
    ap.add_argument("--text", nargs="+", required=True, help="text files (pdftotext output)")
    ap.add_argument("--xlsx", nargs="+", required=True, help="source-data XLSX files")
    ap.add_argument("--value-tol", type=float, default=0.01,
                    help="value proximity tolerance, fraction (default 0.01 = 1 percent)")
    ap.add_argument("--min-value", type=float, default=10.0,
                    help="skip text values below this (default 10; avoids small-int coincidences)")
    ap.add_argument("--bilingual-map", help="optional JSON {english_token: [chinese_tokens]} for cross-lang entity match")
    ap.add_argument("--top-k", type=int, default=2,
                    help="per text claim, report at most this many closest matches (default 2). "
                         "Reduces noise from same-value coincidences across many entities.")
    args = ap.parse_args()

    bilingual = {}
    if args.bilingual_map:
        bilingual = json.loads(Path(args.bilingual_map).read_text())

    # Load XLSX entries
    all_entries = []
    for xp in args.xlsx:
        all_entries.extend(load_xlsx(xp))
    sys.stderr.write(f"loaded {len(all_entries)} numeric cells from {len(args.xlsx)} XLSX file(s)\n")

    # Index by value (rounded to allow proximity search) and by entity token
    value_index: dict[float, list[dict]] = collections.defaultdict(list)
    entity_index: dict[str, list[dict]] = collections.defaultdict(list)
    for e in all_entries:
        value_index[round(e["value"], 2)].append(e)
        for tok in e["entity_tokens"]:
            entity_index[tok].append(e)

    flag_count = 0
    for tpath in args.text:
        text = Path(tpath).read_text(errors="ignore")
        claims = extract_text_claims(text)
        sys.stderr.write(f"{tpath}: {len(claims)} numeric claims\n")
        seen_flags = set()  # (text_offset, xlsx_key)
        for claim in claims:
            if claim["value"] < args.min_value:
                continue
            # Build candidate XLSX entries via entity-token overlap (incl. bilingual)
            candidate_keys = set()
            for tok in claim["context_tokens"]:
                for e in entity_index.get(tok, []):
                    candidate_keys.add(id(e))
                for ch_tok in bilingual.get(tok, []):
                    for e in entity_index.get(ch_tok, []):
                        candidate_keys.add(id(e))
            entity_candidates = [e for e in all_entries if id(e) in candidate_keys]
            # If text has an explicit unit, ALSO scan all XLSX cells whose header has an
            # explicit different unit in the same family — this catches unit-confusion
            # even when entity names don't overlap (e.g., English transliteration in
            # text vs Chinese name in XLSX). The signal is: text-value and xlsx-value
            # are the SAME number to within 1%, but the unit labels claim they should
            # differ by a power of 10. That means one of the unit labels is wrong.
            extra: list[dict] = []
            if claim["unit"] and unit_to_canonical(claim["unit"]) is not None:
                for hits in value_index.values():
                    for e in hits:
                        if not e["header_unit"] or e["value"] == 0:
                            continue
                        if not same_unit_family(claim["unit"], e["header_unit"]):
                            continue
                        if claim["unit"] == e["header_unit"]:
                            continue
                        ratio = claim["value"] / e["value"]
                        # Numbers match literally (within tol) -> one of the unit
                        # labels must be wrong, since the units claim a factor difference.
                        if abs(ratio - 1.0) < args.value_tol:
                            extra.append(e)
            # Collect all candidate hits with their proximity scores, then keep only
            # the top-K closest per text claim. This drops "same-value coincidence"
            # false positives where one text value matches many unrelated entities.
            entity_candidate_ids = {id(c) for c in entity_candidates}
            hits = []  # list of (proximity, hit_type, e, scale_factor)
            seen_e = set()
            for e in entity_candidates + extra:
                if id(e) in seen_e or e["value"] == 0:
                    continue
                seen_e.add(id(e))
                has_entity_overlap = id(e) in entity_candidate_ids
                ratio = claim["value"] / e["value"]
                # Check 1: explicit unit mismatch — same literal value, units claim different scale
                if claim["unit"] and e["header_unit"] and same_unit_family(claim["unit"], e["header_unit"]):
                    cu = unit_to_canonical(claim["unit"])
                    eu = unit_to_canonical(e["header_unit"])
                    if cu and eu and cu != eu and abs(ratio - 1.0) < args.value_tol:
                        prox = abs(ratio - 1.0)
                        hits.append((prox, "unit_confusion", e, cu / eu, has_entity_overlap))
                        continue
                # Check 2: power-of-10 value match. Skip if BOTH sides have explicit but
                # incompatible families (e.g., kV in text vs TWh in XLSX). Different families
                # cannot be a transcription scale-mismatch — they are different physical
                # quantities. This rules out the '800 kV ≠ 79.75 TWh × 10' false-positive class.
                claim_fam = unit_family(claim["unit"]) if claim["unit"] else None
                xlsx_fam = unit_family(e["header_unit"]) if e["header_unit"] else None
                if claim_fam and xlsx_fam and claim_fam != xlsx_fam:
                    continue  # incompatible families, skip Check 2
                for factor in (10.0, 100.0, 1000.0, 1/10, 1/100, 1/1000):
                    if abs(ratio / factor - 1.0) < args.value_tol:
                        prox = abs(ratio / factor - 1.0)
                        hits.append((prox, "scale_mismatch", e, factor, has_entity_overlap))
                        break
            # Rank: entity-overlap hits first (typically high-confidence), then by proximity
            hits.sort(key=lambda h: (not h[4], h[0]))
            for prox, htype, e, scale_factor, has_entity_overlap in hits[: args.top_k]:
                key = (claim["value"], e["sheet"], e["coord"])
                if key in seen_flags:
                    continue
                seen_flags.add(key)
                flag_count += 1
                eo_tag = " (entity-overlap)" if has_entity_overlap else ""
                if htype == "unit_confusion":
                    print(f"\n*** UNIT-CONFUSION FLAG ({claim['unit']} vs {e['header_unit']}, implied factor {scale_factor:g}x){eo_tag} ***")
                    print(f"  text claim: {claim['value']!r} {claim['unit']}")
                else:
                    print(f"\n*** SCALE-MISMATCH FLAG ({scale_factor:g}x){eo_tag} ***")
                    print(f"  text claim: {claim['value']!r} (unit-near: {claim['unit']!r})")
                print(f"  xlsx cell:  {e['value']:.4f} (header={e['header']!r})  in {e['file']}:{e['sheet']}!{e['coord']}")
                print(f"  entity:     {e['entity_raw']!r}  proximity={prox:.5f}")
                print(f"  context:    ...{claim['context'][:200]}...")

    sys.stderr.write(f"\n=== {flag_count} flag(s) emitted ===\n")
    return 0 if flag_count == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
