"""decimal_match.py — cross-cell last-N-decimal matching sweeper.

A deterministic, single-purpose utility script. No LLM SDK. No "skeleton -> enrich"
two-stage orchestration. Use directly from the agent during the numerical-evidence
track when the audit needs to detect cross-cell decimal-trail fabrication.

Usage:
    python decimal_match.py <run-dir>/supplementary/*.xlsx [--digits 2] [--min-class 4] [--ratio 3.0] [--cross-sheet]

What it does:
    For each XLSX sheet, collects every numeric value whose NATURAL precision
    has at least N fractional digits. Groups values by their last N digits,
    then reports groups that exceed uniform-null expectation by a ratio threshold.

    Under H0 (independent biological measurements, last-N-decimals uniform on
    {0..10^N-1}), the expected group size is n_eligible / 10^N. Groups whose
    observed size is >= ratio * expected are flagged.

Why this exists (empirical baseline):
    Kang Tiebang RAB22A osteosarcoma paper (10.1038/s41556-020-0522-z) had a
    whistleblower-reported pattern: "64 data points in ED Fig 6B and 6C have
    identical last two decimals". The skill's v1.4 had no sweeper for this class.
    The deterministic-column-pair sweep (Check 1.1) detects within-row
    additive relationships; the row-duplicate sweep (Check 1.2) detects
    multi-numeric row copies; this sweeper detects something orthogonal:
    cross-cell decimal-trail homogeneity, which arises when many values were
    produced from a small set of seeds + arithmetic.

Limitations (real, document them):
    - Integer-only values are excluded (no fractional part).
    - Values stored as float but at instrument-quantized precision (e.g., a
      flow-cytometer that always reports 2 dp; if the natural quantization is
      already coarse, the sweeper may over-flag). Decide per-column with the
      column-precision check (02-numerical-evidence.md Check 1.4).
    - The uniform-null assumption is approximate for some measurement types
      (counts, ratios near 0 or 1). Use ratio >= 3.0 as a conservative floor.
    - Cross-sheet mode pools across sheets in one workbook; cross-workbook
      pooling is not supported (the relevant comparison is within one
      authoring session, which typically means one workbook).

This script is intentionally < 200 lines. If a paper needs deeper
forensics (e.g., last-N-decimal joint distribution across paired columns),
drop a more specialized script next to this one rather than extending this one.
"""

from __future__ import annotations

import argparse
import collections
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write("install openpyxl first: pip install openpyxl\n")
    sys.exit(2)


def displayed_last_n_decimals(v, n: int, display_dp: int, strict: bool = True) -> str | None:
    """Return last n digits of v at the given display precision, or None.

    Skips values whose float storage carries more precision than display_dp
    (these are typically computed values like 0.6666... that are not at the
    display precision the user reads as "2 decimal places"). Skips integers.

    If strict=True, the value must round-trip exactly at display_dp AND its
    natural precision must equal display_dp (rules out 1.1 -> '1.10' style
    false positives). If strict=False, any value whose natural precision is
    <= display_dp is included.
    """
    if not isinstance(v, (int, float)) or isinstance(v, bool):
        return None
    if v == round(v, 0):  # integer-equivalent
        return None
    if abs(round(v, display_dp) - v) > 1e-9:
        return None  # natural precision > display_dp -> float-noise or true higher precision
    if strict:
        # require natural precision EXACTLY display_dp (not 1.1 displayed at 2dp)
        if abs(round(v, display_dp - 1) - v) < 1e-9:
            return None
    s = f"{round(v, display_dp):.{display_dp}f}"
    frac = s.split(".")[1]
    return frac[-n:]


def sweep_sheet(
    ws, n: int, display_dp: int, strict: bool,
) -> tuple[dict[str, list[tuple[str, float]]], int]:
    """Return (signature -> [(coord, value), ...], total_eligible_cells)."""
    classes: dict[str, list[tuple[str, float]]] = collections.defaultdict(list)
    total = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            sig = displayed_last_n_decimals(cell.value, n, display_dp, strict)
            if sig is None:
                continue
            classes[sig].append((cell.coordinate, cell.value))
            total += 1
    return dict(classes), total


def report_classes(
    title: str,
    classes: dict[str, list],
    total: int,
    n: int,
    min_class: int,
    min_distinct: int,
    ratio: float,
    max_show: int = 12,
) -> int:
    """Print a per-context report. Returns the count of FLAGGED classes.

    A class is FLAGGED only when (cells >= min_class) AND
    (distinct values >= min_distinct) AND (cells/expected >= ratio).
    The distinct-value gate rules out literal cell duplication, which is
    a different fabrication pattern caught by Check 1.2 row-duplicate.
    """
    sig_space = 10 ** n
    expected = total / sig_space if sig_space else 0
    print(f"\n## {title}")
    print(f"   eligible-cells={total}  signature-space=10^{n}={sig_space}  "
          f"uniform-exp-per-class={expected:.3f}")
    flagged = 0
    big = sorted(
        ((sig, cells) for sig, cells in classes.items() if len(cells) >= min_class),
        key=lambda kv: -len(kv[1]),
    )
    if not big:
        print("   (no class meets min-class threshold; sweep clean at these params)")
        return 0
    for sig, cells in big:
        obs = len(cells)
        n_distinct = len({v for _, v in cells})
        r = obs / expected if expected > 0 else float("inf")
        flag = ""
        if n_distinct < min_distinct:
            flag = "  (value-dup, not decimal-match — skip)"
        elif r >= ratio and obs >= min_class:
            flag = "  *** FLAG ***"
            flagged += 1
        print(f"   sig={sig!r:>6}  obs={obs:>4}  distinct={n_distinct:>3}  "
              f"ratio={r:>5.1f}x{flag}")
        if "FLAG" in flag:
            shown = 0
            seen_values = set()
            for coord, v in cells:
                if v in seen_values:
                    continue
                seen_values.add(v)
                print(f"       {coord}: {v}")
                shown += 1
                if shown >= max_show:
                    break
            if n_distinct > max_show:
                print(f"       ... +{n_distinct - max_show} more distinct values")
    return flagged


def main() -> int:
    parser = argparse.ArgumentParser(description=(__doc__ or "").split("\n\n")[0])
    parser.add_argument("files", nargs="+", help="xlsx files to sweep")
    parser.add_argument(
        "--digits", type=int, default=2,
        help="trailing decimal digits to match (default 2)",
    )
    parser.add_argument(
        "--display-dp", type=int, default=None,
        help="display precision of the data (default = --digits)",
    )
    parser.add_argument(
        "--lax", action="store_true",
        help="allow values whose natural precision is < display-dp "
             "(e.g., count 1.1 as '1.10' -> sig '10'). default strict.",
    )
    parser.add_argument(
        "--min-class", type=int, default=4,
        help="only report classes with >= this many cells (default 4)",
    )
    parser.add_argument(
        "--min-distinct", type=int, default=3,
        help="require >= this many distinct values within a class to FLAG "
             "(default 3; rules out literal value duplication)",
    )
    parser.add_argument(
        "--ratio", type=float, default=3.0,
        help="flag classes whose obs/expected >= this ratio (default 3.0)",
    )
    parser.add_argument(
        "--cross-sheet", action="store_true",
        help="also sweep across all sheets in each workbook",
    )
    args = parser.parse_args()
    display_dp = args.display_dp if args.display_dp is not None else args.digits
    strict = not args.lax

    grand_flagged = 0
    for fpath in args.files:
        p = Path(fpath)
        if not p.is_file():
            sys.stderr.write(f"not a file: {fpath}\n")
            continue
        try:
            wb = openpyxl.load_workbook(p, data_only=True)
        except Exception as e:
            sys.stderr.write(f"failed to open {fpath}: {e}\n")
            continue

        # Per-sheet sweep
        workbook_classes: dict[str, list[tuple[str, str, float]]] = collections.defaultdict(list)
        workbook_total = 0
        for sn in wb.sheetnames:
            ws = wb[sn]
            classes, total = sweep_sheet(ws, args.digits, display_dp, strict)
            grand_flagged += report_classes(
                f"{p.name} :: sheet {sn!r}", classes, total,
                args.digits, args.min_class, args.min_distinct, args.ratio,
            )
            if args.cross_sheet:
                for sig, cells in classes.items():
                    workbook_classes[sig].extend((sn, c, v) for c, v in cells)
                workbook_total += total

        if args.cross_sheet and workbook_total > 0:
            sig_space = 10 ** args.digits
            expected = workbook_total / sig_space
            print(f"\n## {p.name} :: CROSS-SHEET pool")
            print(f"   eligible-cells={workbook_total}  uniform-exp-per-class={expected:.3f}")
            big = sorted(
                ((s, items) for s, items in workbook_classes.items() if len(items) >= args.min_class),
                key=lambda kv: -len(kv[1]),
            )
            for sig, items in big:
                sheets_in = {sn for sn, _, _ in items}
                if len(sheets_in) < 2:
                    continue  # within-sheet effects already reported above
                obs = len(items)
                n_distinct = len({v for _, _, v in items})
                r = obs / expected if expected > 0 else float("inf")
                if r < args.ratio or n_distinct < args.min_distinct:
                    continue
                grand_flagged += 1
                print(f"   sig={sig!r:>6}  obs={obs:>4}  distinct={n_distinct:>3}  "
                      f"ratio={r:>5.1f}x  across-sheets={len(sheets_in)}  *** FLAG ***")
                shown = 0
                seen = set()
                for sn, coord, v in items:
                    if v in seen:
                        continue
                    seen.add(v)
                    print(f"       [{sn}] {coord}: {v}")
                    shown += 1
                    if shown >= 12:
                        break
                if n_distinct > 12:
                    print(f"       ... +{n_distinct - 12} more distinct values")

    print(f"\n=== total FLAG count: {grand_flagged} ===")
    return 0 if grand_flagged == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
